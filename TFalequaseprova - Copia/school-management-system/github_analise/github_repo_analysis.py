import requests
import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime
import os
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage

def get_user_repos(username, token=None):
    headers = {}
    if token:
        headers['Authorization'] = f'token {token}'
    url = f'https://api.github.com/users/{username}/repos'
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        return response.json()
    else:
        print(f"Error fetching repositories: {response.status_code}")
        print(response.text)
        return []

def extract_repo_data(repos):
    repo_data = []
    for repo in repos:
        data = {
            'Name': repo['name'],
            'Description': repo['description'] or 'No description',
            'Language': repo['language'] or 'Not specified',
            'Stars': repo['stargazers_count'],
            'Forks': repo['forks_count'],
            'Created At': datetime.strptime(repo['created_at'], '%Y-%m-%dT%H:%M:%SZ'),
            'Last Updated': datetime.strptime(repo['updated_at'], '%Y-%m-%dT%H:%M:%SZ'),
            'URL': repo['html_url']
        }
        repo_data.append(data)
    return repo_data

def save_to_excel(repo_data, username, output_file='github_repos.xlsx'):
    if not repo_data:
        print("No data to save")
        return

    df = pd.DataFrame(repo_data)
    df['Created At'] = pd.to_datetime(df['Created At'])
    df['Last Updated'] = pd.to_datetime(df['Last Updated'])
    df['Created At'] = df['Created At'].dt.strftime('%Y-%m-%d')
    df['Last Updated'] = df['Last Updated'].dt.strftime('%Y-%m-%d')
    df['Days Since Update'] = (datetime.now() - pd.to_datetime(df['Last Updated'])).dt.days

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=f'{username} Repositories', index=False)
        workbook = writer.book
        worksheet = writer.sheets[f'{username} Repositories']
        worksheet.auto_filter.ref = worksheet.dimensions

        red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        worksheet.conditional_formatting.add(
            f'D2:D{len(df)+1}',
            CellIsRule(operator='greaterThan', formula=['10'], fill=red_fill)
        )

        worksheet.cell(row=1, column=len(df.columns)+2, value=f"Data coletada: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

        chart_files = [
            'charts/top_repos_by_stars.png',
            'charts/language_distribution.png',
            'charts/repo_creation_timeline.png'
        ]
        img_row = len(df) + 5
        for chart_path in chart_files:
            if os.path.exists(chart_path):
                img = OpenpyxlImage(chart_path)
                img.anchor = f'A{img_row}'
                worksheet.add_image(img)
                img_row += 30

    print(f"Data and charts saved to {output_file}")
    return df

def generate_analytics(df, username):
    if df.empty:
        print("No data for analysis")
        return

    os.makedirs('charts', exist_ok=True)
    top_repos = df.sort_values('Stars', ascending=False).head(5)
    plt.figure(figsize=(10, 6))
    plt.bar(top_repos['Name'], top_repos['Stars'])
    plt.title(f'Top 5 Repositories by Stars for {username}')
    plt.xlabel('Repository')
    plt.ylabel('Stars')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    plt.savefig('charts/top_repos_by_stars.png')

    language_counts = df['Language'].value_counts()
    plt.figure(figsize=(10, 6))
    language_counts.plot(kind='pie', autopct='%1.1f%%')
    plt.title(f'Programming Languages Used by {username}')
    plt.ylabel('')
    plt.tight_layout()
    plt.savefig('charts/language_distribution.png')

    df['Created At'] = pd.to_datetime(df['Created At'])
    df = df.sort_values('Created At')
    plt.figure(figsize=(12, 6))
    plt.plot(df['Created At'], range(1, len(df) + 1), marker='o')
    plt.title(f'Repository Creation Timeline for {username}')
    plt.xlabel('Date')
    plt.ylabel('Cumulative Number of Repositories')
    plt.grid(True)
    plt.tight_layout()
    plt.savefig('charts/repo_creation_timeline.png')

    print("Analytics generated and saved to 'charts' directory")

def main():
    username = input("Enter GitHub username: ")
    token = input("Enter GitHub token (optional, press Enter to skip): ").strip() or None
    print(f"Fetching repositories for {username}...")
    repos = get_user_repos(username, token)
    if not repos:
        print("No repositories found or error occurred.")
        return
    print(f"Found {len(repos)} repositories.")
    repo_data = extract_repo_data(repos)
    df = save_to_excel(repo_data, username)
    generate_analytics(df, username)
    print("\nSummary:")
    print(f"Total repositories: {len(repos)}")
    if not df.empty:
        print(f"Most starred repository: {df.loc[df['Stars'].idxmax(), 'Name']} ({df['Stars'].max()} stars)")
        print(f"Most forked repository: {df.loc[df['Forks'].idxmax(), 'Name']} ({df['Forks'].max()} forks)")
        print(f"Most used language: {df['Language'].value_counts().index[0]}")

if __name__ == "__main__":
    main()
